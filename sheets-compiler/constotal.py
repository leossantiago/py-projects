# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 15:02:31 2019

@author: leoss
"""

#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
 
import openpyxl
 
#Prepare the spreadsheets to copy from and paste too.
 

import glob
#glob.glob('**/*.xlsx', recursive=True)





 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange():
    rangeSelected = []
    #Loops through selected Rows
    c_lamp = 0
    c_cons = 0.0
    c_ac = 0
    ac_cons = 0.0

    #File to be copied    
    wb = openpyxl.load_workbook(f, data_only=True) #Add file name
    sheet = wb["ILUMINAÇÃO"] #Add Sheet name

    #File to be pasted into
    template = openpyxl.load_workbook("total.xlsx") #Add file name
    temp_sheet = template["PMT"] #Add Sheet name
    
    startCol = 4
    startRow = 11
    endCol = 4
    endRow = sheet.max_row-3


    for i in range(startRow,endRow + 1,1):
        c_lamp = c_lamp + sheet.cell(row = i, column = 4).value
        a = sheet.cell(row = i, column = 6).value
        c_cons = c_cons + a


    #File to be copied
    wb = openpyxl.load_workbook(f, data_only=True) #Add file name
    sheet = wb["CONDICIONADORES DE AR"] #Add Sheet name

    #File to be pasted into
    template = openpyxl.load_workbook("ac.xlsx") #Add file name
    temp_sheet = template["PMT"] #Add Sheet name    
    
    startCol = 6
    startRow = 11
    endCol = 6
    endRow = sheet.max_row-3


    for i in range(startRow,endRow + 1,1):
        c_ac = c_ac + sheet.cell(row = i, column = 6).value
        a = sheet.cell(row = i, column = 8).value
        ac_cons = ac_cons + a


    rowSelected = []
    rowSelected.append(c_lamp)
    rowSelected.append(c_cons)
    rowSelected.append(c_ac)
    rowSelected.append(ac_cons)
    #Adds the RowSelected List and nests inside the rangeSelected
    rangeSelected.append(rowSelected)
    rangeSelected.append([])
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData,source,type):
    

    
    a1 = sheet['A4'].value
    a2 = sheet['A5'].value
    a3 = sheet['A6'].value
    if type == 1:
        a4 = sheet['D1'].value    
    else:
        a4 = sheet['F1'].value
        
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        
        sheetReceiving.cell(row = i, column = 1).value = a1
        sheetReceiving.cell(row = i, column = 2).value = a2
        sheetReceiving.cell(row = i, column = 3).value = a3
        sheetReceiving.cell(row = i, column = 4).value = a4
        
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
        
    
    
    
    
    
path = 'files\\'

files = [f for f in glob.glob(path + "**/*.xlsx", recursive=True)]
countfiles = 1
for f in files:
    #print(f)
    print(countfiles)
    print(f)
    countfiles += 1

#File to be copied    
    wb = openpyxl.load_workbook(f, data_only=True) #Add file name
    sheet = wb["ILUMINAÇÃO"] #Add Sheet name

    #File to be pasted into
    template = openpyxl.load_workbook("total.xlsx") #Add file name
    temp_sheet = template["PMT"] #Add Sheet name



    selectedRange = copyRange()

    pasteRange(5,temp_sheet.max_row+1,8,temp_sheet.max_row+1,temp_sheet,selectedRange,sheet,1)
    
    template.save("total.xlsx")