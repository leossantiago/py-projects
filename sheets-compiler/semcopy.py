# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 15:02:31 2019

@author: leoss
"""

#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
 
import openpyxl
 
#Prepare the spreadsheets to copy from and paste too.
 

import glob
#glob.glob('**/*.xlsx', recursive=True)





 
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData,source,type):
    
    
    
    a1 = sheet['A4'].value
    a2 = sheet['A5'].value
    a3 = sheet['A6'].value
    if type == 1:
        a4 = sheet['D1'].value    
    else:
        a4 = sheet['F1'].value
        
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        
        sheetReceiving.cell(row = i, column = 3).value = a4
        sheetReceiving.cell(row = i, column = 4).value = a1
        sheetReceiving.cell(row = i, column = 5).value = a2
        sheetReceiving.cell(row = i, column = 6).value = a3
        
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
        
    
def createDataIL():
    #print("Processing IL...")
    selectedRange = copyRange(1,11,5,sheet.max_row-3,sheet) #Change the 4 number values
    pasteRange(7,temp_sheet.max_row+1,11,temp_sheet.max_row+sheet.max_row-13,temp_sheet,selectedRange,sheet,1) #Change the 4 number values
    #You can save the template as another file to create a new file here too.
    template.save("ilum.xlsx")
    #print("Range copied and pasted!")
    
def createDataAC():
    #print("Processing AC...")
    selectedRange = copyRange(1,11,7,sheet.max_row-3,sheet) #Change the 4 number values
    pasteRange(7,temp_sheet.max_row+1,13,temp_sheet.max_row+sheet.max_row-13,temp_sheet,selectedRange,sheet,2) #Change the 4 number values
    #You can save the template as another file to create a new file here too.
    template.save("ac.xlsx")
    #print("Range copied and pasted!")
    
    
    
    
    
path = 'files\\'

files = [f for f in glob.glob(path + "**/*.xlsx", recursive=True)]
countfiles = 1
for f in files:
    #print(f)
    print(countfiles)
    countfiles += 1
    #File to be copied
    wb = openpyxl.load_workbook(f) #Add file name
    sheet = wb["ILUMINAÇÃO"] #Add Sheet name

    #File to be pasted into
    template = openpyxl.load_workbook("ilum.xlsx") #Add file name
    temp_sheet = template["PMT"] #Add Sheet name

    go = createDataIL()

    #File to be copied
    wb = openpyxl.load_workbook(f) #Add file name
    sheet = wb["CONDICIONADORES DE AR"] #Add Sheet name

    #File to be pasted into
    template = openpyxl.load_workbook("ac.xlsx") #Add file name
    temp_sheet = template["PMT"] #Add Sheet name

    go = createDataAC()